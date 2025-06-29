import wx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from outputResults import outputAllGradeResults, outputTeamsResults
from getCompetitionNames import getCompetitionNames

# Функция для перевода чисел в римскую систему
def to_roman(number: int) -> str:
    roman_numbers = {'M': 1000, 'CM': 900, 'D': 500, 'CD': 400,
                     'C': 100, 'XC': 90, 'L': 50, 'XL': 40,
                     'X': 10, 'IX': 9, 'V': 5, 'IV': 4, 'I': 1}
    roman = ''
    for letter, value in roman_numbers.items():
        while number >= value:
            roman += letter
            number -= value
    return roman

# Фунция для перевода ступени в возраст
def category_to_age(category: int) -> str:
    res = ''
    match category:
        case 2:
            res = '8-9'
        case 3:
            res = '10-11'
        case 4:
            res = '12-13'
        case 5:
            res = '14-15'
        case 6:
            res = '16-17'
        case 7:
            res = '18-19'
        case 8:
            res = '20-24'
        case 9:
            res = '25-29'
        case 10:
            res = '30-34'
        case 11:
            res = '35-39'
        case 12:
            res = '40-44'
        case 13:
            res = '45-49'
        case 14:
            res = '50-54'
        case 15:
            res = '55-59'
        case 16:
            res = '59-64'
        case 17:
            res = '65-69'
        case 18:
            res = '70+'
    return res

# Класс, хранящий в себе данные о выводимом объекте
class Out_object():
    def __init__(self, place, surname, name, thirdname, date_of_borth: str, number, team, normatives: list,
                 sum) -> None:
        self.name = name
        self.surname = surname
        self.thirdname = thirdname
        self.normatives = normatives
        self.sum = sum
        self.number = number
        self.team = team
        self.place = place
        self.date = date_of_borth

    def get_normatives(self) -> list:
        return self.normatives

    def get_name(self):
        return self.name

    def get_surname(self):
        return self.surname

    def get_thirdname(self):
        return self.thirdname

    def get_sum(self):
        return self.sum

    def get_number(self):
        return self.number

    def get_team(self):
        return self.team

    def get_place(self):
        return self.place

    def get_date(self):
        return self.date

# Класс, хранящий в себе данные о выводимомой ступени
class Out_category():
    def __init__(self, out_objects: list, category: int, sex: str) -> None:
        self.out_objects = out_objects
        self.category = category
        self.sex = sex

    def get_info(self):
        return self.out_objects

    def get_sex(self):
        return self.sex

    def get_category(self):
        return self.category

    def get_normatives(self):
        return self.out_objects[0].get_normatives()

    def get_normatives_without_results(self):
        return [self.out_objects[0].get_normatives()[i] for i in range(0, len(self.out_objects[0].get_normatives()), 3)]

"""Класс, отвечающий за страницу просмотра участников
ОПИСАНИЕ:
    Класс создаёт все виджеты на странице, отвечает за подтягивание данных с базы данных,
    Также отвечает за вывод ошибок в случае неправильной работы с приложением
"""
class Viewer(wx.Panel):
    def __init__(self, parent):

        # Объекты для вывода
        self.out_objects = []
        self.out_categories = []

        # вызов конструктора наследника с указанием того-же родителя, что и у класса Team_viewer
        super(Viewer, self).__init__(parent)
        
        # Задание размера и стандартного шрифта
        self.SetSize((600, 800))
        self.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, 0, "Segoe UI"))
        
        # Создание разметки верхнего уровня        
        self.grid_sizer_1 = wx.FlexGridSizer(11, 1, 0, 0)

        # Создание объекта дял выбора ступени
        self.choice_1 = wx.Choice(self, wx.ID_ANY,
                                  choices=[u"2 ступень", u"3 ступень", u"4 ступень", u"5 ступень", u"6 ступень",
                                           u"7 ступень", u"8 ступень", u"9 ступень", u"10 ступень", u"11 ступень",
                                           u"12 ступень", u"13 ступень", u"14 ступень", u"15 ступень", u"16 ступень",
                                           u"17 ступень", u"18 ступень"])
        self.choice_1.SetSelection(0)
        self.grid_sizer_1.Add(self.choice_1, 0, 0, 0)

        # Создание объекта дял выбора пола
        self.radio_box_1 = wx.RadioBox(self, wx.ID_ANY, "", choices=[u"Мужской", u"Женский"], majorDimension=1,
                                       style=wx.RA_SPECIFY_ROWS)
        self.radio_box_1.SetSelection(0)
        self.grid_sizer_1.Add(self.radio_box_1, 0, 0, 0)
        
        # Создание объекта дял установки начальнх доступных соревнований
        self.normatives = list(getCompetitionNames(2, "Мужской").keys())
        
        # Создание кнопок выбора соревнований
        self.check_list_box_1 = wx.CheckListBox(self, wx.ID_ANY, choices=self.normatives)
        self.check_list_box_1.SetCheckedItems([i for i in range(len(self.normatives))])
        self.grid_sizer_1.Add(self.check_list_box_1, 0, 0, 0)

        # Создание кнопки для выбора/сброса всех соревнований
        self.check_box_1 = wx.CheckBox(self, wx.ID_ANY, 'Включить/выключить всё')
        self.check_box_1.SetValue(True)
        self.grid_sizer_1.Add(self.check_box_1, 0, 0, 0)

        # Добавка кнопки Добавить в протокол
        self.button_2 = wx.Button(self, wx.ID_ANY, "Добавить в протокол")
        self.button_2.SetBackgroundColour(wx.Colour(204, 50, 50))
        self.button_2.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNHIGHLIGHT))
        self.button_2.SetMinSize((350, 25))
        self.grid_sizer_1.Add(self.button_2, 0, 0, 0)

        # Создание кнопки Удалить последнюю запись из протокола
        self.button_4 = wx.Button(self, wx.ID_ANY, "Удалить последнюю запись из протокола")
        self.button_4.SetBackgroundColour(wx.Colour(204, 50, 50))
        self.button_4.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNHIGHLIGHT))
        self.button_4.SetMinSize((350, 25))
        self.grid_sizer_1.Add(self.button_4, 0, 0, 0)

        # Создание кнопки Посмотреть ступень
        self.button_3 = wx.Button(self, wx.ID_ANY, "Посмотреть ступень")
        self.button_3.SetBackgroundColour(wx.Colour(204, 50, 50))
        self.button_3.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNHIGHLIGHT))
        self.button_3.SetMinSize((350, 25))
        self.grid_sizer_1.Add(self.button_3, 0, 0, 0)

        # Создание поля для просмотра полученных данных
        self.list_ctrl_1 = wx.ListCtrl(self, wx.ID_ANY, style=wx.LC_HRULES | wx.LC_REPORT | wx.LC_VRULES)
        self.list_ctrl_1.SetMinSize(wx.Size(10000, 350))
        self.list_ctrl_1.SetFont(wx.Font(12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, 0, ""))
        self.list_ctrl_1.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNHIGHLIGHT))

        self.grid_sizer_1.Add(self.list_ctrl_1, 1, wx.EXPAND, 0)
        # Кнопка Сохранить протокол
        self.button_1 = wx.Button(self, wx.ID_ANY, "Сохранить протокол")
        self.button_1.SetBackgroundColour(wx.Colour(204, 50, 50))
        self.button_1.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNHIGHLIGHT))
        self.button_1.SetMinSize((350, 25))
        self.grid_sizer_1.Add(self.button_1, 0, 0, 0)

        # кнопка Скопировать данные ступени
        self.button_5 = wx.Button(self, wx.ID_ANY, "Скопировать данные ступени")
        self.button_5.SetBackgroundColour(wx.Colour(204, 50, 50))
        self.button_5.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_BTNHIGHLIGHT))
        self.button_5.SetMinSize((350, 25))
        self.grid_sizer_1.Add(self.button_5, 0, 0, 0)

        # Привязка событий
        self.button_1.Bind(wx.EVT_BUTTON, self.save_as_excel)
        self.button_2.Bind(wx.EVT_BUTTON, self.get_data_protocol)
        self.button_3.Bind(wx.EVT_BUTTON, self.get_data)
        self.choice_1.Bind(wx.EVT_CHOICE, self.update_checkListBox_API)
        self.radio_box_1.Bind(wx.EVT_RADIOBOX, self.update_checkListBox_API)
        self.button_4.Bind(wx.EVT_BUTTON, self.delete_last_record)
        self.check_box_1.Bind(wx.EVT_CHECKBOX, self.on_off_all)
        self.button_5.Bind(wx.EVT_BUTTON, self.copy_to_clipboard)
        self.SetSizer(self.grid_sizer_1)

        self.Layout()

    # Функция для отображения данных о протоколе
    def builder_protocol(self):
        if (len(self.out_categories) == 0):
            wx.MessageBox("Нет данных по данной ступени", "Сообщение", wx.OK | wx.ICON_INFORMATION)
            self.list_ctrl_1.DeleteAllColumns()
            self.list_ctrl_1.DeleteAllItems()
            return
        try:
            self.list_ctrl_1.AppendColumn("Ступень в протоколе", format=wx.LIST_FORMAT_LEFT, width=500)
            self.list_ctrl_1.AppendColumn("Пол", format=wx.LIST_FORMAT_LEFT, width=500)
            for i in self.out_categories:
                self.list_ctrl_1.Append([str(i.get_category()), str(i.get_sex())])
        except:
            wx.MessageBox("Сбой программы", "Ошибка", wx.OK | wx.ICON_ERROR)
            self.list_ctrl_1.DeleteAllColumns()
            self.list_ctrl_1.DeleteAllItems()
    
    # Функция для отображения данных о ступени
    def builder(self):
        try:
            self.list_ctrl_1.AppendColumn("Место", format=wx.LIST_FORMAT_LEFT, width=100)
            self.list_ctrl_1.AppendColumn("ФИО", format=wx.LIST_FORMAT_LEFT, width=200)
            self.list_ctrl_1.AppendColumn("Дата рождения", format=wx.LIST_FORMAT_LEFT, width=200)
            self.list_ctrl_1.AppendColumn("Нагрудн. Номер", format=wx.LIST_FORMAT_LEFT, width=200)
            self.list_ctrl_1.AppendColumn("Команда", format=wx.LIST_FORMAT_LEFT, width=200)
            for j in range(0, len(self.out_objects[0].get_normatives()), 3):
                self.list_ctrl_1.AppendColumn(self.out_objects[0].get_normatives()[j], format=wx.LIST_FORMAT_LEFT,
                                              width=200)
                self.list_ctrl_1.AppendColumn("Балл", format=wx.LIST_FORMAT_LEFT, width=200)

            self.list_ctrl_1.AppendColumn("Сумма очков", format=wx.LIST_FORMAT_LEFT, width=100)

            out_category = self.out_categories[len(self.out_categories) - 1]
            for i in range(len(out_category.get_info())):
                out = []
                out.append(str(out_category.get_info()[i].get_place()))
                out.append(
                    out_category.get_info()[i].get_surname() + " " + out_category.get_info()[i].get_name() + " " +
                    out_category.get_info()[i].get_thirdname())
                out.append(out_category.get_info()[i].get_date())
                out.append(out_category.get_info()[i].get_number())
                out.append(out_category.get_info()[i].get_team())
                for j in range(0, len(out_category.get_info()[i].get_normatives())):
                    if (j == 0 or j % 3 == 0):
                        continue
                    out.append(str(out_category.get_info()[i].get_normatives()[j]))
                out.append(str(out_category.get_info()[i].get_sum()))

                self.list_ctrl_1.Append(out)

            self.out_categories.pop()
        except:
            wx.MessageBox("Нет данных по данной ступени", "Сообщение", wx.OK | wx.ICON_INFORMATION)
            self.list_ctrl_1.DeleteAllColumns()
            self.list_ctrl_1.DeleteAllItems()

    # Функция для сохранения данных в эксель
    def save_as_excel(self, event):
        try:
            # Создаем временный Excel файл
            wb = openpyxl.Workbook()
            ws = wb.create_sheet("Личный зачёт")
            del wb['Sheet']
            # Получаем количество строк и колонок
            col_count = 0
            for out_category in self.out_categories:
                temp_col_count = 6
                for i in out_category.get_normatives_without_results():
                    temp_col_count += 2
                if temp_col_count > col_count:
                    col_count = temp_col_count
            cur_row = 1

            # Вывод верхней шапки
            ws.append(["Фестиваль Всероссийского физкультурно-спортивного комплекса \"Готов к труду и обороне\""])
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=col_count)
            cur_row += 1
            ws.append(["Протокол соревнования"])
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=col_count)
            cur_row += 1
            ws.append(["Личный зачёт"])
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=col_count)
            cur_row += 1
            temp = ["<Введите Дату>"]
            for i in range(1, int(col_count / 2)):
                temp.append("")
            temp.append("<Введите Город>")
            ws.append(temp)
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=int(col_count / 2))
            ws.merge_cells(start_row=cur_row, start_column=int(col_count) / 2 + 1, end_row=cur_row,
                           end_column=col_count)
            cur_row += 1

            # Собираем заголовки колонок
            for out_category in self.out_categories:
                normatives_headers = []
                for i in out_category.get_normatives_without_results():
                    normatives_headers.append(i + " Результат")
                    normatives_headers.append("Баллы")
                headers = ["Место", "ФИО", "Дата рождения", "Нагрудн. Номер", "Команда"] + normatives_headers + [
                    "Сумма очков"]
                ws.append(headers)
                # Увеличение размера заголовков
                ws.row_dimensions[cur_row].height = 70

                cur_row += 1
                ws.append([
                              f'({to_roman(out_category.get_category())} ступень)  {category_to_age(out_category.get_category())} лет {out_category.get_sex()}'])
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=col_count)

                # Подсветка цветом категории и возраста
                ws[f'A{cur_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
                cur_row += 1

                # Собираем данные
                for i in range(len(out_category.get_info())):
                    out = []
                    out.append(str(out_category.get_info()[i].get_place()))
                    out.append(
                        out_category.get_info()[i].get_surname() + " " + out_category.get_info()[i].get_name() + " " +
                        out_category.get_info()[i].get_thirdname())
                    out.append(out_category.get_info()[i].get_date())
                    out.append(out_category.get_info()[i].get_number())
                    out.append(out_category.get_info()[i].get_team())
                    for j in range(0, len(out_category.get_info()[i].get_normatives())):
                        if (j == 0 or j % 3 == 0):
                            continue
                        out.append(str(out_category.get_info()[i].get_normatives()[j]))
                    out.append(str(out_category.get_info()[i].get_sum()))

                    ws.append(out)
                    cur_row += 1

            # Добавление нижней части
            temp = ["Главный судья, судья <> категории"]
            for i in range(1, int(col_count / 2)):
                temp.append("")
            temp.append("<Введите ФИО>")
            ws.append(temp)
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=int(col_count / 2))
            ws.merge_cells(start_row=cur_row, start_column=int(col_count) / 2 + 1, end_row=cur_row,
                           end_column=col_count)
            cur_row += 1
            temp = ["Главный секретарь, судья <> категории"]
            for i in range(1, int(col_count / 2)):
                temp.append("")
            temp.append("<Введите ФИО>")
            ws.append(temp)
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=int(col_count / 2))
            ws.merge_cells(start_row=cur_row, start_column=int(col_count) / 2 + 1, end_row=cur_row,
                           end_column=col_count)
           
            # читаемый шрифт
            font_style = Font(name='Times New Roman', size=16, vertAlign='baseline')
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font_style
                    cell.alignment = alignment

            # Верхняя часть
            ws['A1'].font = Font(name='Times New Roman', size=28, vertAlign='baseline', bold=True)
            ws['A2'].font = Font(name='Times New Roman', size=28, vertAlign='baseline', bold=True)
            ws['A3'].font = Font(name='Times New Roman', size=28, vertAlign='baseline', bold=True)
            ws['A4'].font = Font(name='Times New Roman', size=28, vertAlign='baseline')
            ws['G4'].font = Font(name='Times New Roman', size=28, vertAlign='baseline')

            # Размеры ячеек
            ws.row_dimensions[1].height = 70
            ws.row_dimensions[2].height = 40
            ws.row_dimensions[3].height = 40
            ws.row_dimensions[4].height = 40

            for col in range(1, 1000):
                col_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[col_letter].width = 25
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['E'].width = 70

            ws2 = wb.create_sheet("командный зачёт")
            cur_row = 1
            col_count = 3
            ws2.append(["Фестиваль Всероссийского физкультурно-спортивного комплекса \"Готов к труду и обороне\""])
            ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=col_count)
            cur_row += 1
            ws2.append(["Протокол команднного зачёта"])
            ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=col_count)
            cur_row += 1
            temp = ["<Дата>"]
            for i in range(1, int(col_count / 2)):
                temp.append("")
            temp.append("<Город>")
            ws2.append(temp)
            ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=int(col_count / 2))
            ws2.merge_cells(start_row=cur_row, start_column=int(col_count) / 2 + 1, end_row=cur_row,
                            end_column=col_count)
            cur_row += 1

            ws2.row_dimensions[1].height = 80
            ws2.row_dimensions[2].height = 40
            ws2.row_dimensions[3].height = 40
            ws2.row_dimensions[4].height = 40
            ws2.column_dimensions['A'].width = 70
            ws2.column_dimensions['B'].width = 70
            ws2.column_dimensions['C'].width = 70

            headers = ["Место", "Команда", "Сумма очков"]
            ws2.append(headers)
            ws2[f'A{cur_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
            ws2[f'B{cur_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
            ws2[f'C{cur_row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
            cur_row += 1
            example_team_out = outputTeamsResults()

            for i in example_team_out:
                ws2.append(i)
                cur_row += 1

            temp = ["Главный судья, судья <> категории"]
            for i in range(1, int(col_count / 2)):
                temp.append("")
            temp.append("<ФИО>")
            ws2.append(temp)
            ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=int(col_count / 2))
            ws2.merge_cells(start_row=cur_row, start_column=int(col_count) / 2 + 1, end_row=cur_row,
                            end_column=col_count)
            cur_row += 1
            temp = ["Главный секретарь, судья <> категории"]
            for i in range(1, int(col_count / 2)):
                temp.append("")
            temp.append("<ФИО>")

            ws2.append(temp)
            ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=int(col_count / 2))
            ws2.merge_cells(start_row=cur_row, start_column=int(col_count) / 2 + 1, end_row=cur_row,
                            end_column=col_count)
            for row in ws2.iter_rows():
                for cell in row:
                    cell.font = font_style
                    cell.alignment = alignment
            ws2['A1'].font = Font(name='Times New Roman', size=28, vertAlign='baseline', bold=True)
            ws2['A2'].font = Font(name='Times New Roman', size=28, vertAlign='baseline', bold=True)
            ws2['A3'].font = Font(name='Times New Roman', size=28, vertAlign='baseline')
            ws2['B3'].font = Font(name='Times New Roman', size=28, vertAlign='baseline')
            for cell in ws._cells.values():
                cell.border = Border(top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'),
                                     left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'))
            for cell in ws2._cells.values():
                cell.border = Border(top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'),
                                     left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'))
            wb.save('../Итоговый протокол.xlsx')

            # Уведомляем пользователя
            wx.MessageBox("Протокол сохранен в файл \"Итоговый протокол\", в папке gto_helper!", "Успех", wx.OK | wx.ICON_INFORMATION)
        except PermissionError:
            wx.MessageBox("Файл уже открыт, закройте его", "Ошибка", wx.OK | wx.ICON_ERROR)

        except:
            wx.MessageBox("В протоколе нет записей, добавьте хотябы одну", "Ошибка", wx.OK | wx.ICON_ERROR)
            self.list_ctrl_1.DeleteAllColumns()
            self.list_ctrl_1.DeleteAllItems()

    # Функция обработчик нажатия на кнопку добавить в протокол
    def get_data_protocol(self, event):
        try:
            self.create_out_object()
            # Строительство новых данных
            self.builder_protocol()
        except:
            wx.MessageBox("Произошла ошибка", "Неизвестная ошибка", wx.OK | wx.ICON_ERROR)

    # Функция обработчик нажатия на кнопку посмотреть ступень
    def get_data(self, event):
        try:
            self.create_out_object()
            self.builder()
        except:
            wx.MessageBox("Произошла ошибка", "Неизвестная ошибка", wx.OK | wx.ICON_ERROR)

    # Дробавление в внтуренние данные информации о добавленом объекте
    def create_out_object(self):
        # Удаление старых записей
        self.list_ctrl_1.DeleteAllColumns()
        self.list_ctrl_1.DeleteAllItems()
        # Удаление соревнований, которые не выбраны в программе
        white_list = self.check_list_box_1.GetCheckedStrings()

        # Для запроса в базу
        category = self.choice_1.GetSelection() + 2
        sex = self.radio_box_1.GetString(self.radio_box_1.GetSelection())
        self.out_objects.clear()

        # Пример формата получаемых данных
        for i in outputAllGradeResults(category, sex):
            filtred_norrmatives = []
            for j in range(0, len(i[7]), 3):
                if i[7][j] in white_list:
                    filtred_norrmatives.append(i[7][j])
                    filtred_norrmatives.append(i[7][j + 1])
                    filtred_norrmatives.append(i[7][j + 2])

            self.out_objects.append(Out_object(i[0], i[1], i[2], i[3], i[4], i[5], i[6], filtred_norrmatives, i[8]))
        temp_out_object = [i for i in self.out_objects]
        if (len(temp_out_object) > 0): self.out_categories.append(Out_category(temp_out_object, category, sex))

    #  Функция обновляющая список соревнований
    def update_checkListBox(self):
        category = self.choice_1.GetSelection() + 2
        sex = self.radio_box_1.GetString(self.radio_box_1.GetSelection())
        self.normatives = list(getCompetitionNames(category, sex).keys())
        self.check_list_box_1.SetItems(self.normatives)
        self.check_list_box_1.SetCheckedItems([i for i in range(len(self.normatives))])
        self.check_box_1.SetValue(True)
    
    def update_checkListBox_API(self, event):
        self.update_checkListBox()

    #  Функция удаляющая последнюю запись из протокола
    def delete_last_record(self, event):
        try:
            self.out_categories.pop()
            wx.MessageBox("Удаление успешно", "Успех", wx.ICON_INFORMATION | wx.OK)
            # Удаление старых записей
            self.list_ctrl_1.DeleteAllColumns()
            self.list_ctrl_1.DeleteAllItems()
            self.builder_protocol()
        except:
            wx.MessageBox("В протоколе нет записей", "Информация", wx.ICON_INFORMATION | wx.OK)

    # Функция выбирающая/сбрасываю0щая выбор с соревнований
    def on_off_all(self, event):
        category = self.choice_1.GetSelection() + 2
        sex = self.radio_box_1.GetString(self.radio_box_1.GetSelection())
        self.normatives = list(getCompetitionNames(category, sex).keys())
        if self.check_box_1.GetValue():
            self.check_list_box_1.SetCheckedItems([i for i in range(len(self.normatives))])
        else:
            self.check_list_box_1.SetCheckedItems([])
    # Функция копипрования ступени в буфер обмена
    def copy_to_clipboard(self, event):
        row_count = self.list_ctrl_1.GetItemCount()
        col_count = self.list_ctrl_1.GetColumnCount()
        all_items = []
        # Копируем в буфер обмена
        for row in range(row_count):
            row_data = []
            for col in range(col_count):
                row_data.append(self.list_ctrl_1.GetItemText(row, col))
            all_items.append("\t".join(row_data))
        if all_items:
            clipboard_data = wx.TextDataObject()
            clipboard_data.SetText("\n".join(all_items))
            if wx.TheClipboard.Open():
                wx.TheClipboard.SetData(clipboard_data)
                wx.TheClipboard.Close()
        wx.MessageBox("Данные скопированы в буфер обмена!", "Успех", wx.OK | wx.ICON_INFORMATION)
